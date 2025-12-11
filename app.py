import os
import secrets
import logging
import requests
from pathlib import Path
from decimal import Decimal
from datetime import datetime, date
from flask import Flask, render_template, request, session, jsonify
from openpyxl import load_workbook
import db

# Load environment variables from .env file
db.load_env_from_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def is_xlsx_filename(filename: str) -> bool:
    return filename.lower().endswith(".xlsx")

def decimals_from_format(fmt: str) -> int:
    if ".000" in fmt:
        return 3
    if ".00" in fmt:
        return 2
    if ".0" in fmt:
        return 1
    return 0

def format_numeric(value, fmt: str) -> str:
    d = decimals_from_format(fmt)
    use_grouping = "," in fmt or "#" in fmt
    negative_parentheses = "(" in fmt and ")" in fmt
    currency_symbol = "$" if ("$" in fmt or "[$" in fmt) else ""
    if negative_parentheses and (isinstance(value, (int, float, Decimal)) and value < 0):
        v = abs(value)
        formatted = f"{v:,.{d}f}" if use_grouping else f"{v:.{d}f}"
        return f"({currency_symbol}{formatted})"
    formatted = f"{value:,.{d}f}" if use_grouping else f"{value:.{d}f}"
    return f"{currency_symbol}{formatted}" if currency_symbol else formatted

def format_cell_display(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    fmt = cell.number_format or ""
    if isinstance(v, (int, float, Decimal)):
        if any(s in fmt for s in ["$", "[$", "0", "#", "%"]):
            if "%" in fmt:
                d = decimals_from_format(fmt)
                return f"{v*100:.{d}f}%"
            return format_numeric(v, fmt)
        return str(v)
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    return str(v)

def read_excel_display(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [format_cell_display(c) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2):
        rows.append([format_cell_display(c) for c in row])
    return headers, rows

def read_excel_raw(filepath):
    """
    Reads Excel and returns raw data for DB insertion.
    Expected columns: SITEID, COSTCENTER, NAME, LITERSLOADED, PRICE, DATE, IMPORT
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = []
    
    # Iterate over rows, starting from row 2
    for row in ws.iter_rows(min_row=2):
        r = []
        # We only care about first 7 columns
        for i, cell in enumerate(row):
            if i > 6:
                break
            
            val = cell.value
            
            # Logic for numeric rounding to match display (Columns: LITERS=3, PRICE=4, IMPORT=6)
            if i in [3, 4, 6] and isinstance(val, (int, float, Decimal)):
                fmt = cell.number_format
                if fmt and fmt.lower() != 'general':
                    if ".0" in fmt:
                        d = decimals_from_format(fmt)
                        val = round(float(val), d)
                    elif "0" in fmt or "#" in fmt:
                        # Integer format implied
                        val = round(float(val), 0)
            
            # Logic for DATE (index 5)
            if i == 5:
                if isinstance(val, datetime):
                    val = val.date()
                elif isinstance(val, str):
                    try:
                        val = datetime.strptime(val, "%d/%m/%Y").date()
                    except ValueError:
                        try:
                            val = datetime.strptime(val, "%Y-%m-%d").date()
                        except ValueError:
                            pass
                            
            r.append(val)
        
        # Pad if row is short
        while len(r) < 7:
            r.append(None)
            
        rows.append(r)
        
    return rows

@app.route("/", methods=["GET"]) 
def index():
    return render_template("index.html", error_message=None, headers=[], rows=[])

@app.route("/upload", methods=["POST"]) 
def upload():
    file = request.files.get("file")
    if not file or file.filename == "":
        return render_template("index.html", error_message="Selecciona un archivo .xlsx.", headers=[], rows=[])
    if not is_xlsx_filename(file.filename):
        return render_template("index.html", error_message="Solo se permiten archivos .xlsx.", headers=[], rows=[])
    
    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)
    session['uploaded_file'] = filepath
    
    headers, rows = read_excel_display(filepath)
    return render_template("index.html", error_message=None, headers=headers, rows=rows)

@app.route("/send", methods=["POST"])
def send_data():
    logging.info("Recibida petición para enviar datos a HANA.")
    filepath = session.get('uploaded_file')
    if not filepath or not os.path.exists(filepath):
        logging.warning("Intento de envío fallido: No hay archivo cargado o sesión expirada.")
        return jsonify({"success": False, "message": "No hay archivo cargado o expiró la sesión."}), 400
    
    try:
        logging.info(f"Procesando archivo: {filepath}")
        raw_data = read_excel_raw(filepath)
        logging.info(f"Archivo leído correctamente. Filas extraídas: {len(raw_data)}")
        
        logging.info("Iniciando inserción en base de datos...")
        db.insert_gas_data(raw_data)
        logging.info("Inserción en HANA completada exitosamente.")
        
        # Llamada al servicio externo para procesar la información
        service_url = "https://snbrns-processes-hub-noisy-baboon-ll.cfapps.us10.hana.ondemand.com/snbrns-hub/hana/procedures/sp-snbrs-19"
        logging.info(f"Iniciando llamada al servicio externo: {service_url}")
        
        try:
            # Usamos POST ya que es un trigger de proceso. 
            # Se envían parámetros por defecto requeridos por el servicio
            payload = {
                "param1": 0,
                "param2": "trigger_automatico"
            }
            logging.info(f"Enviando payload al servicio: {payload}")
            response = requests.post(service_url, json=payload, timeout=300) 
            
            if response.status_code == 200 or response.status_code == 201:
                logging.info("Llamada al servicio completada exitosamente.")
                try:
                    resp_json = response.json()
                    logging.info(f"Respuesta del servicio: {resp_json}")
                    
                    # Validamos el campo 'success' si existe en la respuesta
                    if resp_json and isinstance(resp_json, dict):
                         if not resp_json.get('success', True):
                             error_msg = resp_json.get('message', 'El servicio externo reportó un error.')
                             logging.warning(f"El servicio respondió con éxito HTTP pero con success=false: {error_msg}")
                             raise Exception(f"Error en procesamiento externo: {error_msg}")
                except ValueError:
                    # Si no es JSON, solo registramos que no se pudo parsear pero seguimos si el status es 200
                    logging.warning("La respuesta del servicio no es un JSON válido, pero el código de estado es correcto.")
            else:
                logging.error(f"Error en el servicio externo. Status: {response.status_code}. Respuesta: {response.text}")
                raise Exception(f"El servicio de procesamiento falló con código {response.status_code}.")
                
        except requests.exceptions.RequestException as req_err:
            logging.error(f"Excepción de conexión con el servicio externo: {str(req_err)}")
            raise Exception(f"Error de conexión con el servicio de procesamiento: {str(req_err)}")

        return jsonify({"success": True, "message": "Datos enviados y procesados correctamente."})
    except Exception as e:
        logging.error(f"Error crítico al enviar datos: {str(e)}", exc_info=True)
        return jsonify({"success": False, "message": f"Error al enviar datos: {str(e)}"}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
